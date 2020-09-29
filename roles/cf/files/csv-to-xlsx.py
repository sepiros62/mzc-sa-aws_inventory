#-*- coding:utf-8 -*-
import os
import csv
import sys

from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font

bd = Side(style='thin', color="000000")
default_border = Border(left=bd, top=bd, right=bd, bottom=bd)
gray_fill = PatternFill("solid", fgColor="00C0C0C0")

bottom_line = Border(bottom=Side(border_style="medium", color='00FF0000'))

ft_head_line = Font(name="Malgun Gothic", size="24", bold=True)
ft_mid_line = Font(name="Malgun Gothic", size="11", bold=True)
ft_bot_line = Font(name="Malgun Gothic", size="10", bold=True)
ft_con_line = Font(name="Arial", size="7", bold=False)

center_alignment = Alignment(horizontal="center", vertical="center")

def excel_initialized(sheet_name):
    ws = sheet_name
    ws.row_dimensions[1].height = 16
    ws.row_dimensions[2].height = 15
    ws.row_dimensions[3].height = 15
    ws.row_dimensions[4].height = 15
    ws.row_dimensions[5].height = 15
    ws.row_dimensions[6].height = 15
    ws.row_dimensions[7].height = 15
    ws.row_dimensions[8].height = 15

    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 60
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 13
    ws.column_dimensions['E'].width = 13
    ws.column_dimensions['F'].width = 24
    ws.column_dimensions['G'].width = 13
    ws.column_dimensions['H'].width = 16

    title_cell = ws.cell(row=2, column=2, value="Megazone Infra Monthly Report")
    title_cell.font = ft_head_line
    title_cell.border = bottom_line

#    ws['b6'].font = ft_bot_line
#    ws['b6'].fill = gray_fill
#    ws.merge_cells('B2:J2')

def private_cell_merge(worksheets, sr, er, sc, ec):
    worksheets.merge_cells(start_row=sr, end_row=er, start_column=sc, end_column=ec)
    return ec + 1

def subject_style(sheetname):
    sheetname.font = ft_bot_line
    sheetname.fill = gray_fill
    sheetname.alignment = center_alignment
    sheetname.border = default_border

def con_style(sheetname):
    sheetname.alignment = center_alignment
    sheetname.border = default_border
    sheetname.font = ft_con_line


# Main 
wb = Workbook()
ws = wb.active
ws.title="test"

excel_initialized(ws)

with open('cf-template.csv', 'r') as f:
    reader = csv.reader(f)
    for r, row in enumerate(reader):
        for c, col in enumerate(row):
            for idx, val in enumerate(col.split(',')):
              cell = ws.cell(row=r+1, column=c+1)
              cell.value = val
    wb.save('cf-template.xlsx')
